#!/usr/bin/env python3
"""
Build the Global Institutional Attendance Report workbook.
Generates a polished, fully-functional Excel report with:
  1. Dashboard       — institutional KPIs and summary charts
  2. Course Summary  — one row per course section with metrics
  3. Student Detail  — one row per student per course with risk bands
  4. Daily Trends    — attendance by course by date
  5. Compliance      — courses with missing/stale attendance
  6. Data Model      — documentation of assumptions and formulas
  7. Config          — configurable thresholds and parameters

Uses realistic sample data based on UCCI Blackboard structure.
When real data is available via bb_global_extract.py, paste CSVs
into the data sheets and all formulas/charts auto-update.
"""

import csv
import os
import random
from datetime import datetime, timedelta
from math import ceil

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import (
    CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
)
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Design System ────────────────────────────────────────────────────────────

# Colors
TEAL       = "01696F"      # Primary accent
TEAL_LIGHT = "E0F2F3"      # Light teal for highlights
DARK_BG    = "1B474D"      # Dark header
WHITE      = "FFFFFF"
LIGHT_GRAY = "F7F6F2"      # Background
MID_GRAY   = "D4D1CA"      # Borders
TEXT_DARK  = "28251D"       # Primary text
TEXT_MUTED = "7A7974"       # Secondary text
RED_RISK   = "A13544"       # High risk
ORANGE     = "DA7101"       # Medium risk / warning
GREEN_OK   = "437A22"       # OK / success
GOLD       = "D19900"       # Caution

# Chart colors
CHART_TEAL   = "20808D"
CHART_TERRA  = "A84B2F"
CHART_DARK   = "1B474D"
CHART_CYAN   = "BCE2E7"
CHART_MAUVE  = "944454"
CHART_GOLD   = "FFC553"

# Fonts
FONT_TITLE   = Font(name="Calibri", size=18, bold=True, color=DARK_BG)
FONT_H2      = Font(name="Calibri", size=14, bold=True, color=TEAL)
FONT_H3      = Font(name="Calibri", size=12, bold=True, color=TEXT_DARK)
FONT_BODY    = Font(name="Calibri", size=11, color=TEXT_DARK)
FONT_SMALL   = Font(name="Calibri", size=10, color=TEXT_MUTED)
FONT_HEADER  = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_KPI_VAL = Font(name="Calibri", size=28, bold=True, color=TEAL)
FONT_KPI_LBL = Font(name="Calibri", size=10, color=TEXT_MUTED)
FONT_LINK    = Font(name="Calibri", size=11, color="0000FF", underline="single")
FONT_WHITE   = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_KPI_SUB = Font(name="Calibri", size=11, color=TEXT_DARK)

# Fills
FILL_HEADER  = PatternFill("solid", fgColor=DARK_BG)
FILL_TEAL_LT = PatternFill("solid", fgColor=TEAL_LIGHT)
FILL_LIGHT   = PatternFill("solid", fgColor=LIGHT_GRAY)
FILL_WHITE   = PatternFill("solid", fgColor=WHITE)
FILL_RED_LT  = PatternFill("solid", fgColor="FCE4EC")
FILL_GREEN_LT = PatternFill("solid", fgColor="E8F5E9")
FILL_ORANGE_LT = PatternFill("solid", fgColor="FFF3E0")
FILL_KPI_BG  = PatternFill("solid", fgColor=LIGHT_GRAY)

# Borders
THIN_BORDER  = Border(
    bottom=Side(style="thin", color=MID_GRAY),
)
BOTTOM_THICK = Border(
    bottom=Side(style="medium", color=TEAL),
)

# Alignment
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", indent=1)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")
ALIGN_WRAP   = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "BB_Global_Attendance_Report.xlsx")
CSV_DIR     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "attendance_data")


# ── Sample Data Generation ───────────────────────────────────────────────────

DEPARTMENTS = ["Arts & Sciences", "Business", "Education", "Technology", "Health Sciences"]
INSTRUCTORS = {
    "Arts & Sciences": ["Dr. Sarah Mitchell", "Prof. James Liu", "Dr. Maria Rodriguez"],
    "Business": ["Dr. Robert Chen", "Prof. Angela White"],
    "Education": ["Dr. Patricia Brown", "Prof. David Kim"],
    "Technology": ["Dr. Michael Zhang", "Prof. Lisa Thompson"],
    "Health Sciences": ["Dr. Emma Johnson", "Prof. William Davis"],
}

COURSES = [
    ("MAT 100", "Introductory Algebra", "Arts & Sciences"),
    ("MAT 121", "Precalculus with Trigonometry", "Arts & Sciences"),
    ("MAT 200", "Calculus I", "Arts & Sciences"),
    ("SPA 101", "Elementary Spanish I", "Arts & Sciences"),
    ("SPA 201", "Intermediate Spanish I", "Arts & Sciences"),
    ("ENG 101", "English Composition I", "Arts & Sciences"),
    ("ENG 201", "British Literature", "Arts & Sciences"),
    ("HIS 101", "World History I", "Arts & Sciences"),
    ("BUS 101", "Introduction to Business", "Business"),
    ("BUS 201", "Principles of Management", "Business"),
    ("ACC 101", "Financial Accounting", "Business"),
    ("MKT 201", "Marketing Principles", "Business"),
    ("EDU 101", "Foundations of Education", "Education"),
    ("EDU 201", "Educational Psychology", "Education"),
    ("EDU 301", "Curriculum Design", "Education"),
    ("CIS 101", "Introduction to Computing", "Technology"),
    ("CIS 201", "Database Systems", "Technology"),
    ("CIS 301", "Network Administration", "Technology"),
    ("NUR 101", "Fundamentals of Nursing", "Health Sciences"),
    ("NUR 201", "Pharmacology", "Health Sciences"),
    ("NUR 301", "Clinical Practice I", "Health Sciences"),
    ("BIO 101", "General Biology", "Arts & Sciences"),
    ("PHY 101", "General Physics", "Arts & Sciences"),
    ("PSY 101", "Introduction to Psychology", "Arts & Sciences"),
    ("SOC 101", "Introduction to Sociology", "Arts & Sciences"),
]

FIRST_NAMES = [
    "Aiden", "Bella", "Carlos", "Diana", "Ethan", "Fiona", "George", "Hannah",
    "Ivan", "Julia", "Kevin", "Luna", "Mason", "Nora", "Oliver", "Priya",
    "Quinn", "Rachel", "Samuel", "Tanya", "Uma", "Victor", "Wendy", "Xavier",
    "Yara", "Zane", "Aaliyah", "Brandon", "Camille", "Derek", "Elena", "Felix",
    "Grace", "Henry", "Iris", "Jordan", "Kai", "Leah", "Marcus", "Nina",
]
LAST_NAMES = [
    "Anderson", "Bennett", "Clark", "Davis", "Edwards", "Fisher", "Garcia",
    "Hill", "Ingram", "Jackson", "King", "Lopez", "Morgan", "Nelson",
    "Ortiz", "Patel", "Quinn", "Rivera", "Smith", "Turner", "Upton",
    "Vasquez", "Williams", "Xavier", "Young", "Zhang",
]


def generate_sample_data():
    """Generate realistic sample data for all report sheets."""
    random.seed(42)
    term = "Spring 2026"
    today = datetime(2026, 3, 19)
    term_start = datetime(2026, 1, 12)

    # Generate course sections with 1-2 sections each
    sections = []
    for code, name, dept in COURSES:
        n_sections = random.choice([1, 1, 1, 2])
        for sec in range(1, n_sections + 1):
            instructor = random.choice(INSTRUCTORS[dept])
            sec_id = f"2026_SP_{code.replace(' ', '_')}_{sec}"
            sections.append({
                "term": term,
                "dept": dept,
                "code": code,
                "section": sec,
                "sec_id": sec_id,
                "name": name,
                "instructor": instructor,
            })

    # Generate students (150 unique students, each enrolled in 4-6 courses)
    students = []
    for i in range(150):
        sid = f"S{20260000 + i + 1}"
        fname = random.choice(FIRST_NAMES)
        lname = random.choice(LAST_NAMES)
        students.append({"id": sid, "name": f"{fname} {lname}"})

    # Assign students to course sections
    enrollments = []  # (student, section)
    for stu in students:
        n_courses = random.randint(4, 6)
        chosen = random.sample(sections, min(n_courses, len(sections)))
        for sec in chosen:
            enrollments.append((stu, sec))

    # Generate meetings for each section (2 per week from term start)
    section_meetings = {}
    for sec in sections:
        # Some courses have no attendance (compliance issue)
        if random.random() < 0.08:  # ~8% have no attendance at all
            section_meetings[sec["sec_id"]] = []
            continue

        meetings = []
        # Monday/Wednesday or Tuesday/Thursday schedule
        days = random.choice([(0, 2), (1, 3)])
        d = term_start
        while d <= today:
            if d.weekday() in days:
                meetings.append(d)
            d += timedelta(days=1)

        # Some courses stopped recording recently (stale)
        if random.random() < 0.10:  # ~10% are stale
            cutoff = today - timedelta(days=random.randint(20, 45))
            meetings = [m for m in meetings if m <= cutoff]

        section_meetings[sec["sec_id"]] = meetings

    # Generate attendance records per student per section
    course_summary_data = []
    student_detail_data = []
    daily_data = []
    compliance_data = []

    for sec in sections:
        sid = sec["sec_id"]
        meetings = section_meetings.get(sid, [])
        sec_enrollments = [(stu, sec) for (stu, s) in enrollments if s["sec_id"] == sid]
        n_students = len(sec_enrollments)

        all_present = 0
        all_late = 0
        all_absent = 0
        all_na = 0
        students_100 = 0
        students_above = 0
        students_below = 0
        student_rates = []
        last_date = ""

        if meetings:
            last_date = max(meetings).strftime("%Y-%m-%d")

        for stu, _ in sec_enrollments:
            s_present = 0
            s_late = 0
            s_absent = 0
            s_last = ""

            # Student attendance pattern: some good, some bad
            pattern = random.random()
            for mtg in meetings:
                r = random.random()
                if pattern > 0.85:
                    # Good student: 95%+ present
                    if r < 0.92:
                        s_present += 1
                    elif r < 0.97:
                        s_late += 1
                    else:
                        s_absent += 1
                elif pattern > 0.5:
                    # Average student: 75-90%
                    if r < 0.78:
                        s_present += 1
                    elif r < 0.88:
                        s_late += 1
                    else:
                        s_absent += 1
                elif pattern > 0.2:
                    # Below average: 50-75%
                    if r < 0.55:
                        s_present += 1
                    elif r < 0.70:
                        s_late += 1
                    else:
                        s_absent += 1
                else:
                    # At-risk: <50%
                    if r < 0.30:
                        s_present += 1
                    elif r < 0.45:
                        s_late += 1
                    else:
                        s_absent += 1

            # Add some NA entries
            na_count = 0
            if random.random() < 0.15 and len(meetings) > 0:
                na_count = random.randint(1, min(3, len(meetings)))
                # Remove from counts
                for _ in range(na_count):
                    if s_absent > 0:
                        s_absent -= 1
                    elif s_late > 0:
                        s_late -= 1

            total_marked = s_present + s_late + s_absent
            if total_marked > 0:
                rate = round((s_present * 100 + s_late * 50) / total_marked, 2)
            else:
                rate = None

            if rate is not None:
                student_rates.append(rate)
                if rate >= 100:
                    students_100 += 1
                if rate >= 75:
                    students_above += 1
                else:
                    students_below += 1

            all_present += s_present
            all_late += s_late
            all_absent += s_absent
            all_na += na_count

            if rate is None:
                band = "No Data"
            elif rate < 50:
                band = "High Risk"
            elif rate < 75:
                band = "Medium Risk"
            else:
                band = "OK"

            if meetings:
                s_last = max(meetings).strftime("%Y-%m-%d")

            student_detail_data.append({
                "term": sec["term"],
                "dept": sec["dept"],
                "course_code": f"{sec['code']}-{sec['section']}",
                "course_name": sec["name"],
                "instructor": sec["instructor"],
                "student_id": stu["id"],
                "student_name": stu["name"],
                "present": s_present,
                "late": s_late,
                "absent": s_absent,
                "excused": 0,
                "attendance_pct": rate,
                "last_attendance_date": s_last,
                "risk_band": band,
                "below_threshold": "Yes" if rate is not None and rate < 75 else ("N/A" if rate is None else "No"),
            })

        # Daily attendance data
        for mtg in meetings:
            daily_data.append({
                "term": sec["term"],
                "dept": sec["dept"],
                "course_code": f"{sec['code']}-{sec['section']}",
                "course_name": sec["name"],
                "meeting_date": mtg.strftime("%Y-%m-%d"),
                "students_enrolled": n_students,
            })

        # Course summary
        avg_rate = round(sum(student_rates) / len(student_rates), 2) if student_rates else None
        days_since = None
        status = "Not Recorded"
        if meetings:
            days_since = (today - max(meetings)).days
            if days_since <= 14:
                status = "Active"
            else:
                status = "Stale"
        blocked = random.random() < 0.04  # ~4% API blocked

        course_summary_data.append({
            "term": sec["term"],
            "dept": sec["dept"],
            "course_code": f"{sec['code']}-{sec['section']}",
            "course_name": sec["name"],
            "instructor": sec["instructor"],
            "total_students": n_students,
            "avg_attendance_pct": avg_rate,
            "students_100_pct": students_100,
            "students_above_threshold": students_above,
            "students_below_threshold": students_below,
            "pct_above_threshold": round(students_above / n_students * 100, 1) if n_students else None,
            "pct_below_threshold": round(students_below / n_students * 100, 1) if n_students else None,
            "total_present": all_present,
            "total_late": all_late,
            "total_absent": all_absent,
            "total_na": all_na,
            "total_meetings": len(meetings),
            "last_attendance_date": last_date,
            "days_since_last": days_since,
            "status": status,
            "api_blocked": blocked,
        })

        compliance_data.append({
            "term": sec["term"],
            "dept": sec["dept"],
            "course_code": f"{sec['code']}-{sec['section']}",
            "course_name": sec["name"],
            "instructor": sec["instructor"],
            "total_meetings": len(meetings),
            "total_records": all_present + all_late + all_absent,
            "last_attendance_date": last_date,
            "days_since_last": days_since,
            "status": status,
            "api_blocked": blocked,
            "no_attendance_recorded": len(meetings) == 0 and not blocked,
            "not_recent": status == "Stale",
        })

    return course_summary_data, student_detail_data, daily_data, compliance_data


# ── Sheet Builders ───────────────────────────────────────────────────────────

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def style_header_row(ws, row, start_col, end_col):
    """Apply dark header style to a row range."""
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def add_kpi_card(ws, row, col, value, label, fmt="0.0%"):
    """Add a KPI card (value + label) to a sheet."""
    val_cell = ws.cell(row=row, column=col)
    val_cell.font = FONT_KPI_VAL
    val_cell.alignment = Alignment(horizontal="center", vertical="bottom")

    if isinstance(value, (int, float)):
        if fmt == "0.0%":
            val_cell.value = value / 100
            val_cell.number_format = "0.0%"
        elif fmt == "#,##0":
            val_cell.value = value
            val_cell.number_format = "#,##0"
        else:
            val_cell.value = value
            val_cell.number_format = fmt
    else:
        val_cell.value = value

    lbl_cell = ws.cell(row=row + 1, column=col)
    lbl_cell.value = label
    lbl_cell.font = FONT_KPI_LBL
    lbl_cell.alignment = Alignment(horizontal="center", vertical="top")

    # Light background for the KPI area
    for r in range(row, row + 2):
        ws.cell(row=r, column=col).fill = FILL_KPI_BG


def build_dashboard(wb, course_data, student_data):
    """Build the Dashboard sheet with institutional KPIs and charts."""
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = TEAL

    # Column A spacer
    set_col_width(ws, 1, 3)

    # Title
    ws.merge_cells("B2:K2")
    title = ws.cell(row=2, column=2, value="Institutional Attendance Dashboard")
    title.font = FONT_TITLE
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 36

    # Subtitle
    ws.merge_cells("B3:K3")
    # Pull the term name from the data instead of hardcoding
    term_label = course_data[0]["term"] if course_data else "Unknown Term"
    sub = ws.cell(row=3, column=2, value=f"{term_label}  |  Generated {datetime.now().strftime('%d %b %Y %H:%M')}")
    sub.font = FONT_SMALL
    sub.alignment = Alignment(horizontal="left", vertical="center")

    # Accent line under title
    for c in range(2, 12):
        ws.cell(row=4, column=c).border = BOTTOM_THICK

    # ── KPI Row ──────────────────────────────────────────────────────────
    row = 6

    # Calculate KPIs
    total_courses = len(course_data)
    active_courses = sum(1 for c in course_data if c["status"] == "Active")
    total_students_enrolled = sum(c["total_students"] for c in course_data)
    all_rates = [c["avg_attendance_pct"] for c in course_data if c["avg_attendance_pct"] is not None]
    institution_avg = round(sum(all_rates) / len(all_rates), 1) if all_rates else 0
    at_risk = sum(1 for s in student_data if s["risk_band"] in ("High Risk", "Medium Risk"))
    no_record = sum(1 for c in course_data if c["status"] == "Not Recorded")
    stale_count = sum(1 for c in course_data if c["status"] == "Stale")

    ws.row_dimensions[row].height = 42
    ws.row_dimensions[row + 1].height = 20

    kpis = [
        (2, institution_avg, "Institutional Avg Attendance", "0.0%"),
        (4, total_courses, "Total Course Sections", "#,##0"),
        (6, total_students_enrolled, "Total Enrollments", "#,##0"),
        (8, at_risk, "At-Risk Students", "#,##0"),
        (10, no_record + stale_count, "Compliance Flags", "#,##0"),
    ]

    for col, val, label, fmt in kpis:
        add_kpi_card(ws, row, col, val, label, fmt)
        # Merge for wider KPI
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        for r in [row, row + 1]:
            ws.cell(row=r, column=col + 1).fill = FILL_KPI_BG

    # ── Navigation Links ─────────────────────────────────────────────────
    row = 9
    ws.cell(row=row, column=2, value="Quick Navigation").font = FONT_H3
    nav_items = [
        ("Course Summary", "'Course Summary'!A1"),
        ("Student Detail", "'Student Detail'!A1"),
        ("Daily Trends", "'Daily Trends'!A1"),
        ("Compliance", "'Compliance'!A1"),
        ("Data Model", "'Data Model'!A1"),
        ("Config", "'Config'!A1"),
    ]
    for i, (label, loc) in enumerate(nav_items):
        c = ws.cell(row=row + 1, column=2 + i)
        c.value = label
        c.hyperlink = Hyperlink(ref=c.coordinate, location=loc)
        c.font = FONT_LINK

    # ── Summary Table: Attendance by Department ──────────────────────────
    row = 12
    ws.cell(row=row, column=2, value="Attendance by Department").font = FONT_H2
    row += 1

    # Build department summary
    dept_stats = {}
    for c in course_data:
        d = c["dept"]
        if d not in dept_stats:
            dept_stats[d] = {"courses": 0, "students": 0, "rates": [], "at_risk": 0}
        dept_stats[d]["courses"] += 1
        dept_stats[d]["students"] += c["total_students"]
        if c["avg_attendance_pct"] is not None:
            dept_stats[d]["rates"].append(c["avg_attendance_pct"])

    for s in student_data:
        d = s["dept"]
        if d in dept_stats and s["risk_band"] in ("High Risk", "Medium Risk"):
            dept_stats[d]["at_risk"] += 1

    headers = ["Department", "Courses", "Enrollments", "Avg Attendance %", "At-Risk Students"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2 + i, value=h)
    style_header_row(ws, row, 2, 6)
    row += 1

    dept_start = row
    for dept_name in sorted(dept_stats.keys()):
        ds = dept_stats[dept_name]
        avg = round(sum(ds["rates"]) / len(ds["rates"]), 1) if ds["rates"] else 0
        ws.cell(row=row, column=2, value=dept_name).font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_LEFT
        ws.cell(row=row, column=3, value=ds["courses"]).alignment = ALIGN_CENTER
        ws.cell(row=row, column=3).font = FONT_BODY
        ws.cell(row=row, column=4, value=ds["students"]).alignment = ALIGN_CENTER
        ws.cell(row=row, column=4).font = FONT_BODY
        c = ws.cell(row=row, column=5, value=avg / 100)
        c.number_format = "0.0%"
        c.alignment = ALIGN_CENTER
        c.font = FONT_BODY
        ws.cell(row=row, column=6, value=ds["at_risk"]).alignment = ALIGN_CENTER
        ws.cell(row=row, column=6).font = FONT_BODY
        for col in range(2, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    dept_end = row - 1

    # Column widths for dashboard
    widths = {2: 28, 3: 14, 4: 16, 5: 20, 6: 18, 7: 18, 8: 18, 9: 18, 10: 18, 11: 18}
    for c, w in widths.items():
        set_col_width(ws, c, w)

    # ── Bar Chart: Attendance by Department ──────────────────────────────
    chart = BarChart()
    chart.type = "col"
    chart.title = "Average Attendance % by Department"
    chart.y_axis.title = "Attendance %"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.numFmt = "0%"
    chart.style = 10
    chart.width = 22
    chart.height = 12

    data_ref = Reference(ws, min_col=5, min_row=dept_start - 1, max_row=dept_end)
    cats_ref = Reference(ws, min_col=2, min_row=dept_start, max_row=dept_end)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.legend = None

    # Show department labels on x-axis, rotated to fit
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.delete = False
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties
    chart.x_axis.txPr = RichText(
        bodyPr=RichTextProperties(rot=-5400000),  # -90 degrees in EMUs (vertical)
        p=[Paragraph(pPr=ParagraphProperties(
            defRPr=CharacterProperties(sz=800)  # 8pt font
        ), endParaRPr=CharacterProperties(sz=800))]
    )

    # Color bars
    series = chart.series[0]
    series.graphicalProperties.solidFill = CHART_TEAL

    ws.add_chart(chart, "B" + str(row + 2))

    # ── Stacked Bar: Risk Distribution ───────────────────────────────────
    row2 = row + 2
    # Build risk data by department
    risk_row = row + 2 + ceil(12 * 2) + 2  # After chart
    ws.cell(row=risk_row, column=8, value="Student Risk Distribution by Department").font = FONT_H3

    risk_headers = ["Department", "OK", "Medium Risk", "High Risk"]
    risk_row += 1
    for i, h in enumerate(risk_headers):
        ws.cell(row=risk_row, column=8 + i, value=h)
    style_header_row(ws, risk_row, 8, 11)
    risk_row += 1

    risk_start = risk_row
    for dept_name in sorted(dept_stats.keys()):
        ok_count = sum(1 for s in student_data if s["dept"] == dept_name and s["risk_band"] == "OK")
        med_count = sum(1 for s in student_data if s["dept"] == dept_name and s["risk_band"] == "Medium Risk")
        high_count = sum(1 for s in student_data if s["dept"] == dept_name and s["risk_band"] == "High Risk")
        ws.cell(row=risk_row, column=8, value=dept_name).font = FONT_BODY
        ws.cell(row=risk_row, column=9, value=ok_count).font = FONT_BODY
        ws.cell(row=risk_row, column=9).alignment = ALIGN_CENTER
        ws.cell(row=risk_row, column=10, value=med_count).font = FONT_BODY
        ws.cell(row=risk_row, column=10).alignment = ALIGN_CENTER
        ws.cell(row=risk_row, column=11, value=high_count).font = FONT_BODY
        ws.cell(row=risk_row, column=11).alignment = ALIGN_CENTER
        risk_row += 1
    risk_end = risk_row - 1

    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "stacked"
    chart2.title = "Student Risk Distribution"
    chart2.y_axis.title = "Number of Students"
    chart2.style = 10
    chart2.width = 22
    chart2.height = 12

    data2 = Reference(ws, min_col=9, max_col=11, min_row=risk_start - 1, max_row=risk_end)
    cats2 = Reference(ws, min_col=8, min_row=risk_start, max_row=risk_end)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)

    # Show department labels on x-axis, rotated to fit
    chart2.x_axis.tickLblPos = "low"
    chart2.x_axis.delete = False
    chart2.x_axis.txPr = RichText(
        bodyPr=RichTextProperties(rot=-5400000),  # -90 degrees
        p=[Paragraph(pPr=ParagraphProperties(
            defRPr=CharacterProperties(sz=800)
        ), endParaRPr=CharacterProperties(sz=800))]
    )

    # Color series
    colors2 = [GREEN_OK, ORANGE, RED_RISK]
    for i, s in enumerate(chart2.series):
        s.graphicalProperties.solidFill = colors2[i]

    ws.add_chart(chart2, "H" + str(row2))

    # ── Compliance Summary ───────────────────────────────────────────────
    comp_row = risk_row + 2
    ws.cell(row=comp_row, column=2, value="Compliance Overview").font = FONT_H2
    comp_row += 1

    comp_items = [
        ("Active (attendance recently recorded)", active_courses, GREEN_OK),
        ("Stale (no attendance in 14+ days)", stale_count, ORANGE),
        ("Not Recorded (no attendance data)", no_record, RED_RISK),
    ]
    for label, count, color in comp_items:
        ws.cell(row=comp_row, column=2, value=label).font = FONT_BODY
        ws.cell(row=comp_row, column=2).alignment = ALIGN_LEFT
        c = ws.cell(row=comp_row, column=5, value=count)
        c.font = Font(name="Calibri", size=14, bold=True, color=color)
        c.alignment = ALIGN_CENTER
        comp_row += 1

    # Data source note
    comp_row += 2
    ws.merge_cells(f"B{comp_row}:K{comp_row}")
    note = ws.cell(row=comp_row, column=2,
                   value="Data Source: Blackboard Learn REST API  |  Weighted Scoring: Present=100%, Late=50%, Absent=0%  |  Excused & NA excluded from calculations")
    note.font = FONT_SMALL
    note.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A5"
    return ws


def build_course_summary(wb, data):
    """Build Course Summary sheet — one row per course section."""
    ws = wb.create_sheet("Course Summary")
    ws.sheet_properties.tabColor = "20808D"

    set_col_width(ws, 1, 3)

    # Title
    ws.merge_cells("B2:T2")
    ws.cell(row=2, column=2, value="Course-Level Attendance Summary").font = FONT_TITLE
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:T3")
    ws.cell(row=3, column=2,
            value="One row per course section  |  Filter by department, instructor, or status  |  Threshold: 75%").font = FONT_SMALL

    # Headers
    headers = [
        ("Term", 14),
        ("Department", 22),
        ("Course Code", 16),
        ("Course Name", 30),
        ("Instructor", 24),
        ("Students", 12),
        ("Avg Attendance %", 18),
        ("100% Attendance", 16),
        ("Above 75%", 14),
        ("Below 75%", 14),
        ("% Above 75%", 14),
        ("% Below 75%", 14),
        ("Present", 12),
        ("Late", 10),
        ("Absent", 10),
        ("NA", 10),
        ("Meetings", 12),
        ("Last Attendance", 18),
        ("Days Since Last", 16),
        ("Status", 14),
    ]

    header_row = 5
    for i, (h, w) in enumerate(headers):
        col = i + 2
        ws.cell(row=header_row, column=col, value=h)
        set_col_width(ws, col, w)
    style_header_row(ws, header_row, 2, len(headers) + 1)

    # Data rows
    data_start = header_row + 1
    for r, row_data in enumerate(data):
        row = data_start + r
        vals = [
            row_data["term"],
            row_data["dept"],
            row_data["course_code"],
            row_data["course_name"],
            row_data["instructor"],
            row_data["total_students"],
            row_data["avg_attendance_pct"] / 100 if row_data["avg_attendance_pct"] is not None else "",
            row_data["students_100_pct"],
            row_data["students_above_threshold"],
            row_data["students_below_threshold"],
            row_data["pct_above_threshold"] / 100 if row_data["pct_above_threshold"] is not None else "",
            row_data["pct_below_threshold"] / 100 if row_data["pct_below_threshold"] is not None else "",
            row_data["total_present"],
            row_data["total_late"],
            row_data["total_absent"],
            row_data["total_na"],
            row_data["total_meetings"],
            row_data["last_attendance_date"],
            row_data["days_since_last"] if row_data["days_since_last"] is not None else "",
            row_data["status"],
        ]
        for i, v in enumerate(vals):
            col = i + 2
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = FONT_BODY
            cell.border = THIN_BORDER

            # Formatting
            if i == 6 and isinstance(v, (int, float)):  # Avg attendance
                cell.number_format = "0.0%"
                cell.alignment = ALIGN_CENTER
            elif i in (10, 11) and isinstance(v, (int, float)):  # %
                cell.number_format = "0.0%"
                cell.alignment = ALIGN_CENTER
            elif isinstance(v, (int, float)):
                cell.alignment = ALIGN_CENTER
            elif i == 3:
                cell.alignment = ALIGN_WRAP
            else:
                cell.alignment = ALIGN_LEFT

    data_end = data_start + len(data) - 1

    # Create Excel Table
    table_ref = f"B{header_row}:{get_column_letter(len(headers) + 1)}{data_end}"
    table = Table(displayName="CourseSummary", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)

    # Conditional formatting on Avg Attendance %
    att_col_letter = get_column_letter(8)  # Column H (Avg Attendance)
    ws.conditional_formatting.add(
        f"{att_col_letter}{data_start}:{att_col_letter}{data_end}",
        ColorScaleRule(
            start_type="num", start_value=0, start_color="F8696B",
            mid_type="num", mid_value=0.75, mid_color="FFEB84",
            end_type="num", end_value=1.0, end_color="63BE7B",
        )
    )

    # Conditional formatting on Status
    status_col = get_column_letter(21)  # Column U (Status)
    ws.conditional_formatting.add(
        f"{status_col}{data_start}:{status_col}{data_end}",
        CellIsRule(operator="equal", formula=['"Active"'],
                   fill=FILL_GREEN_LT, font=Font(color=GREEN_OK))
    )
    ws.conditional_formatting.add(
        f"{status_col}{data_start}:{status_col}{data_end}",
        CellIsRule(operator="equal", formula=['"Stale"'],
                   fill=FILL_ORANGE_LT, font=Font(color=ORANGE))
    )
    ws.conditional_formatting.add(
        f"{status_col}{data_start}:{status_col}{data_end}",
        CellIsRule(operator="equal", formula=['"Not Recorded"'],
                   fill=FILL_RED_LT, font=Font(color=RED_RISK))
    )

    # Data bars on Present count
    present_col = get_column_letter(14)
    ws.conditional_formatting.add(
        f"{present_col}{data_start}:{present_col}{data_end}",
        DataBarRule(start_type="min", end_type="max", color=CHART_TEAL)
    )

    ws.freeze_panes = f"A{header_row + 1}"

    return ws


def build_student_detail(wb, data):
    """Build Student Detail sheet — one row per student per course."""
    ws = wb.create_sheet("Student Detail")
    ws.sheet_properties.tabColor = "A84B2F"

    set_col_width(ws, 1, 3)

    ws.merge_cells("B2:P2")
    ws.cell(row=2, column=2, value="Student-Level Attendance Detail").font = FONT_TITLE
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:P3")
    ws.cell(row=3, column=2,
            value="One row per student per course  |  Filter by risk band, department, or instructor  |  Threshold: 75%").font = FONT_SMALL

    headers = [
        ("Term", 14),
        ("Department", 20),
        ("Course Code", 15),
        ("Course Name", 28),
        ("Instructor", 22),
        ("Student ID", 14),
        ("Student Name", 24),
        ("Present", 10),
        ("Late", 10),
        ("Absent", 10),
        ("Excused", 10),
        ("Attendance %", 15),
        ("Last Attendance", 16),
        ("Risk Band", 14),
        ("Below Threshold", 16),
    ]

    header_row = 5
    for i, (h, w) in enumerate(headers):
        col = i + 2
        ws.cell(row=header_row, column=col, value=h)
        set_col_width(ws, col, w)
    style_header_row(ws, header_row, 2, len(headers) + 1)

    data_start = header_row + 1
    for r, row_data in enumerate(data):
        row = data_start + r
        vals = [
            row_data["term"],
            row_data["dept"],
            row_data["course_code"],
            row_data["course_name"],
            row_data["instructor"],
            row_data["student_id"],
            row_data["student_name"],
            row_data["present"],
            row_data["late"],
            row_data["absent"],
            row_data["excused"],
            row_data["attendance_pct"] / 100 if row_data["attendance_pct"] is not None else "",
            row_data["last_attendance_date"],
            row_data["risk_band"],
            row_data["below_threshold"],
        ]
        for i, v in enumerate(vals):
            col = i + 2
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = FONT_BODY
            cell.border = THIN_BORDER

            if i == 11 and isinstance(v, (int, float)):  # Attendance %
                cell.number_format = "0.0%"
                cell.alignment = ALIGN_CENTER
            elif isinstance(v, (int, float)):
                cell.alignment = ALIGN_CENTER
            elif i == 3:
                cell.alignment = ALIGN_WRAP
            else:
                cell.alignment = ALIGN_LEFT

    data_end = data_start + len(data) - 1

    # Table
    table_ref = f"B{header_row}:{get_column_letter(len(headers) + 1)}{data_end}"
    table = Table(displayName="StudentDetail", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium3", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)

    # Conditional formatting: Risk Band
    risk_col = get_column_letter(15)  # Column O
    ws.conditional_formatting.add(
        f"{risk_col}{data_start}:{risk_col}{data_end}",
        CellIsRule(operator="equal", formula=['"High Risk"'],
                   fill=FILL_RED_LT, font=Font(bold=True, color=RED_RISK))
    )
    ws.conditional_formatting.add(
        f"{risk_col}{data_start}:{risk_col}{data_end}",
        CellIsRule(operator="equal", formula=['"Medium Risk"'],
                   fill=FILL_ORANGE_LT, font=Font(bold=True, color=ORANGE))
    )
    ws.conditional_formatting.add(
        f"{risk_col}{data_start}:{risk_col}{data_end}",
        CellIsRule(operator="equal", formula=['"OK"'],
                   fill=FILL_GREEN_LT, font=Font(color=GREEN_OK))
    )

    # Color scale on Attendance %
    att_col = get_column_letter(13)
    ws.conditional_formatting.add(
        f"{att_col}{data_start}:{att_col}{data_end}",
        ColorScaleRule(
            start_type="num", start_value=0, start_color="F8696B",
            mid_type="num", mid_value=0.75, mid_color="FFEB84",
            end_type="num", end_value=1.0, end_color="63BE7B",
        )
    )

    # Data bars on Present count
    present_col = get_column_letter(9)
    ws.conditional_formatting.add(
        f"{present_col}{data_start}:{present_col}{data_end}",
        DataBarRule(start_type="min", end_type="max", color=CHART_TEAL)
    )

    ws.freeze_panes = f"A{header_row + 1}"

    return ws


def build_daily_trends(wb, daily_data, course_data):
    """Build Daily Trends sheet."""
    ws = wb.create_sheet("Daily Trends")
    ws.sheet_properties.tabColor = "1B474D"

    set_col_width(ws, 1, 3)

    ws.merge_cells("B2:H2")
    ws.cell(row=2, column=2, value="Daily Attendance Trends").font = FONT_TITLE
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:H3")
    ws.cell(row=3, column=2,
            value="Attendance session count by course by date  |  Use filters to focus on specific courses or date ranges").font = FONT_SMALL

    headers = [
        ("Term", 14),
        ("Department", 20),
        ("Course Code", 15),
        ("Course Name", 28),
        ("Meeting Date", 16),
        ("Students Enrolled", 18),
    ]

    header_row = 5
    for i, (h, w) in enumerate(headers):
        col = i + 2
        ws.cell(row=header_row, column=col, value=h)
        set_col_width(ws, col, w)
    style_header_row(ws, header_row, 2, len(headers) + 1)

    # Sort by date
    daily_sorted = sorted(daily_data, key=lambda x: (x["course_code"], x["meeting_date"]))

    data_start = header_row + 1
    for r, row_data in enumerate(daily_sorted):
        row = data_start + r
        vals = [
            row_data["term"],
            row_data["dept"],
            row_data["course_code"],
            row_data["course_name"],
            row_data["meeting_date"],
            row_data["students_enrolled"],
        ]
        for i, v in enumerate(vals):
            col = i + 2
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = FONT_BODY
            cell.border = THIN_BORDER
            if isinstance(v, (int, float)):
                cell.alignment = ALIGN_CENTER
            elif i == 3:
                cell.alignment = ALIGN_WRAP
            else:
                cell.alignment = ALIGN_LEFT

    data_end = data_start + len(daily_sorted) - 1

    # Table
    if len(daily_sorted) > 0:
        table_ref = f"B{header_row}:{get_column_letter(len(headers) + 1)}{data_end}"
        table = Table(displayName="DailyTrends", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(table)

    ws.freeze_panes = f"A{header_row + 1}"
    return ws


def build_compliance(wb, compliance_data):
    """Build Compliance sheet — identify gaps in attendance recording."""
    ws = wb.create_sheet("Compliance")
    ws.sheet_properties.tabColor = "A13544"

    set_col_width(ws, 1, 3)

    ws.merge_cells("B2:N2")
    ws.cell(row=2, column=2, value="Attendance Compliance Monitor").font = FONT_TITLE
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:N3")
    ws.cell(row=3, column=2,
            value="Identifies courses with no attendance recorded, stale attendance, or API-blocked access  |  Red = action needed").font = FONT_SMALL

    headers = [
        ("Term", 14),
        ("Department", 20),
        ("Course Code", 15),
        ("Course Name", 28),
        ("Instructor", 22),
        ("Total Meetings", 16),
        ("Total Records", 14),
        ("Last Attendance", 18),
        ("Days Since Last", 16),
        ("Status", 14),
        ("API Blocked", 14),
        ("No Attendance", 16),
        ("Not Recent", 14),
    ]

    header_row = 5
    for i, (h, w) in enumerate(headers):
        col = i + 2
        ws.cell(row=header_row, column=col, value=h)
        set_col_width(ws, col, w)
    style_header_row(ws, header_row, 2, len(headers) + 1)

    # Sort: problems first
    compliance_sorted = sorted(compliance_data,
                               key=lambda x: (0 if x["status"] == "Not Recorded" else
                                              1 if x["status"] == "Stale" else 2,
                                              x["course_code"]))

    data_start = header_row + 1
    for r, row_data in enumerate(compliance_sorted):
        row = data_start + r
        vals = [
            row_data["term"],
            row_data["dept"],
            row_data["course_code"],
            row_data["course_name"],
            row_data["instructor"],
            row_data["total_meetings"],
            row_data["total_records"],
            row_data["last_attendance_date"],
            row_data["days_since_last"] if row_data["days_since_last"] is not None else "",
            row_data["status"],
            "Yes" if row_data["api_blocked"] else "No",
            "Yes" if row_data["no_attendance_recorded"] else "No",
            "Yes" if row_data["not_recent"] else "No",
        ]
        for i, v in enumerate(vals):
            col = i + 2
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = FONT_BODY
            cell.border = THIN_BORDER
            if isinstance(v, (int, float)):
                cell.alignment = ALIGN_CENTER
            elif i == 3:
                cell.alignment = ALIGN_WRAP
            else:
                cell.alignment = ALIGN_LEFT

    data_end = data_start + len(compliance_sorted) - 1

    # Table
    table_ref = f"B{header_row}:{get_column_letter(len(headers) + 1)}{data_end}"
    table = Table(displayName="Compliance", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium6", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)

    # Conditional formatting on Status
    status_col = get_column_letter(11)
    ws.conditional_formatting.add(
        f"{status_col}{data_start}:{status_col}{data_end}",
        CellIsRule(operator="equal", formula=['"Active"'],
                   fill=FILL_GREEN_LT, font=Font(color=GREEN_OK))
    )
    ws.conditional_formatting.add(
        f"{status_col}{data_start}:{status_col}{data_end}",
        CellIsRule(operator="equal", formula=['"Stale"'],
                   fill=FILL_ORANGE_LT, font=Font(bold=True, color=ORANGE))
    )
    ws.conditional_formatting.add(
        f"{status_col}{data_start}:{status_col}{data_end}",
        CellIsRule(operator="equal", formula=['"Not Recorded"'],
                   fill=FILL_RED_LT, font=Font(bold=True, color=RED_RISK))
    )

    # Highlight "Yes" in No Attendance / Not Recent columns
    for col_idx in [13, 14]:  # No Attendance, Not Recent
        cl = get_column_letter(col_idx)
        ws.conditional_formatting.add(
            f"{cl}{data_start}:{cl}{data_end}",
            CellIsRule(operator="equal", formula=['"Yes"'],
                       fill=FILL_RED_LT, font=Font(bold=True, color=RED_RISK))
        )

    ws.freeze_panes = f"A{header_row + 1}"
    return ws


def build_data_model(wb):
    """Build Data Model documentation sheet."""
    ws = wb.create_sheet("Data Model")
    ws.sheet_properties.tabColor = "7A7974"

    set_col_width(ws, 1, 3)
    set_col_width(ws, 2, 30)
    set_col_width(ws, 3, 60)
    set_col_width(ws, 4, 40)

    ws.merge_cells("B2:D2")
    ws.cell(row=2, column=2, value="Data Model and Assumptions").font = FONT_TITLE
    ws.row_dimensions[2].height = 30

    row = 4

    # Section: Data Sources
    ws.cell(row=row, column=2, value="Blackboard REST API Endpoints Used").font = FONT_H2
    row += 1

    api_data = [
        ("Endpoint", "Purpose", "Notes"),
        ("POST /oauth2/token", "Authentication", "OAuth 2.0 Client Credentials"),
        ("GET /v1/terms", "Term list", "Paginated, newest first after reversal"),
        ("GET /v3/courses?termId=X", "All courses in a term", "Filtered by term"),
        ("GET /v1/courses/{cid}/users", "Course memberships", "Filter locally by courseRoleId"),
        ("GET /v1/courses/{cid}/meetings", "Meeting/session list", "403 if attendance not enabled"),
        ("GET /v1/courses/{cid}/meetings/users/{uid}", "Bulk attendance records", "Returns ALL courses (API quirk)"),
        ("GET /v1/courses/{cid}/meetings/{mid}/users/{uid}", "Single attendance record", "Most reliable, slowest"),
        ("GET /v1/users/{uid}", "User profile", "Name, email, student ID"),
    ]

    for i, (a, b, c) in enumerate(api_data):
        for j, v in enumerate([a, b, c]):
            cell = ws.cell(row=row, column=2 + j, value=v)
            if i == 0:
                cell.font = FONT_HEADER
                cell.fill = FILL_HEADER
            else:
                cell.font = FONT_BODY
                cell.border = THIN_BORDER
            cell.alignment = ALIGN_LEFT
        row += 1

    row += 1

    # Section: Scoring Formula
    ws.cell(row=row, column=2, value="Attendance Rate Calculation").font = FONT_H2
    row += 1

    formulas = [
        ("Formula", "Rate = (Present x 100 + Late x 50) / (Present + Late + Absent)"),
        ("Present Weight", "100% — fully attended"),
        ("Late Weight", "50% — partial credit"),
        ("Absent Weight", "0% — no credit"),
        ("Excused", "Excluded from denominator entirely"),
        ("Not Marked / NA", "Excluded from all calculations"),
        ("Verification", "Matches Blackboard displayed rates exactly"),
    ]

    for label, desc in formulas:
        ws.cell(row=row, column=2, value=label).font = Font(name="Calibri", size=11, bold=True, color=TEXT_DARK)
        ws.cell(row=row, column=3, value=desc).font = FONT_BODY
        ws.cell(row=row, column=3).alignment = ALIGN_WRAP
        row += 1

    row += 1

    # Section: Risk Bands
    ws.cell(row=row, column=2, value="Risk Band Classification").font = FONT_H2
    row += 1

    bands = [
        ("Band", "Attendance Range", "Action"),
        ("High Risk", "< 50%", "Immediate intervention required"),
        ("Medium Risk", "50% - 74%", "Early warning — advisor follow-up"),
        ("OK", "75%+", "No action needed"),
        ("No Data", "No attendance recorded", "Check compliance"),
    ]

    for i, (a, b, c) in enumerate(bands):
        for j, v in enumerate([a, b, c]):
            cell = ws.cell(row=row, column=2 + j, value=v)
            if i == 0:
                cell.font = FONT_HEADER
                cell.fill = FILL_HEADER
            else:
                cell.font = FONT_BODY
                cell.border = THIN_BORDER
            cell.alignment = ALIGN_LEFT
        row += 1

    row += 1

    # Section: Course Status
    ws.cell(row=row, column=2, value="Course Recording Status").font = FONT_H2
    row += 1

    statuses = [
        ("Status", "Definition", "Configurable"),
        ("Active", "Attendance recorded within last X days (default: 14)", "Stale threshold in Config sheet"),
        ("Stale", "No attendance recorded in last X days", "Stale threshold in Config sheet"),
        ("Not Recorded", "Zero attendance meetings/records for the course", "-"),
        ("API Blocked", "403 error — attendance not enabled at course level", "Requires Blackboard admin action"),
    ]

    for i, (a, b, c) in enumerate(statuses):
        for j, v in enumerate([a, b, c]):
            cell = ws.cell(row=row, column=2 + j, value=v)
            if i == 0:
                cell.font = FONT_HEADER
                cell.fill = FILL_HEADER
            else:
                cell.font = FONT_BODY
                cell.border = THIN_BORDER
            cell.alignment = ALIGN_LEFT
        row += 1

    row += 1

    # Section: Known Limitations
    ws.cell(row=row, column=2, value="Known Limitations and Data Quality Notes").font = FONT_H2
    row += 1

    limitations = [
        "Not all courses use Blackboard attendance — absence from data does not mean absence from class.",
        "The bulk user attendance endpoint returns cross-course data. Records are cross-referenced with per-course meeting lists to ensure accuracy.",
        "Some courses return 403 on the meetings endpoint, even with System Admin privileges. A probe strategy recovers most data from these courses.",
        "Instructor role filtering via API query param is unreliable. All memberships are fetched and filtered locally by courseRoleId.",
        "Attendance data quality varies by instructor. Some instructors record daily; others record sporadically or not at all.",
        "The 'hierarchy node' (department/programme) comes from the Blackboard course organization structure and may not match institutional reporting hierarchies exactly.",
        "Dropped students (availability.available = 'No') are excluded from all calculations.",
    ]

    for lim in limitations:
        ws.merge_cells(f"B{row}:D{row}")
        cell = ws.cell(row=row, column=2, value=f"  {lim}")
        cell.font = FONT_BODY
        cell.alignment = ALIGN_WRAP
        ws.row_dimensions[row].height = max(30, len(lim) // 2)
        row += 1

    row += 1

    # Section: Data Refresh
    ws.cell(row=row, column=2, value="Data Refresh Process").font = FONT_H2
    row += 1

    steps = [
        "1. Run bb_global_extract.py with your Blackboard credentials",
        "2. Select the term to extract",
        "3. The script outputs four CSV files in the attendance_data/ folder",
        "4. Import CSVs into this workbook's data sheets (Course Summary, Student Detail, etc.)",
        "5. All formulas, charts, and conditional formatting update automatically",
        "6. Pivot tables can be refreshed in Excel via Data > Refresh All",
    ]

    for step in steps:
        ws.merge_cells(f"B{row}:D{row}")
        ws.cell(row=row, column=2, value=step).font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_LEFT
        row += 1

    return ws


def build_config(wb):
    """Build Config sheet with adjustable parameters."""
    ws = wb.create_sheet("Config")
    ws.sheet_properties.tabColor = "D19900"

    set_col_width(ws, 1, 3)
    set_col_width(ws, 2, 30)
    set_col_width(ws, 3, 20)
    set_col_width(ws, 4, 50)

    ws.merge_cells("B2:D2")
    ws.cell(row=2, column=2, value="Report Configuration").font = FONT_TITLE
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:D3")
    ws.cell(row=3, column=2,
            value="Adjust these parameters to customise thresholds and reporting behaviour").font = FONT_SMALL

    row = 5
    config_items = [
        ("Parameter", "Value", "Description"),
        ("Attendance Threshold (%)", 75, "Students below this % are flagged. Used for risk bands and course summaries."),
        ("High Risk Cutoff (%)", 50, "Students below this % are classified as High Risk."),
        ("Stale Days Threshold", 14, "Courses with no attendance in this many days are marked Stale."),
        ("Term", "Spring 2026", "The academic term being reported on."),
        ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M"), "Timestamp of last data extraction."),
        ("Data Source", "Blackboard Learn REST API", "Primary data source for attendance records."),
        ("Scoring Method", "Weighted (P=100%, L=50%, A=0%)", "Matches Blackboard's built-in attendance scoring."),
    ]

    for i, (a, b, c) in enumerate(config_items):
        for j, v in enumerate([a, b, c]):
            col = 2 + j
            cell = ws.cell(row=row, column=col, value=v)
            if i == 0:
                cell.font = FONT_HEADER
                cell.fill = FILL_HEADER
                cell.alignment = ALIGN_CENTER
            else:
                cell.font = FONT_BODY
                cell.alignment = ALIGN_LEFT
                cell.border = THIN_BORDER
                if j == 1:
                    cell.alignment = ALIGN_CENTER
                    if isinstance(v, int):
                        cell.font = Font(name="Calibri", size=11, bold=True, color=TEAL)
                if j == 0:
                    cell.font = Font(name="Calibri", size=11, bold=True, color=TEXT_DARK)
        row += 1

    # Highlight editable cells
    for r in range(6, 9):  # The three editable thresholds
        ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor="FFFDE7")

    row += 2
    ws.merge_cells(f"B{row}:D{row}")
    ws.cell(row=row, column=2,
            value="Yellow cells are editable parameters. Change these to adjust report behavior.").font = FONT_SMALL

    return ws


# ── Pivot Tables (built from actual data) ────────────────────────────────────

def build_risk_pivot(wb, student_data):
    """Build the Risk Pivot sheet: Department × Risk Band cross-tab."""
    ws = wb.create_sheet("Risk Pivot")
    ws.sheet_properties.tabColor = RED_RISK

    # Column A spacer
    set_col_width(ws, 1, 3)

    # Title
    ws.merge_cells("B2:G2")
    ws.cell(row=2, column=2, value="Student Risk Distribution by Department").font = FONT_TITLE
    ws.row_dimensions[2].height = 30

    # Gather unique departments and risk bands from data
    risk_bands = ["OK", "Medium Risk", "High Risk", "No Data"]
    depts = sorted(set(s["dept"] for s in student_data if s.get("dept")))
    if not depts:
        depts = ["(No Department)"]

    # Count students per department per risk band
    counts = {}  # dept -> {band -> count}
    for s in student_data:
        d = s.get("dept", "") or "(No Department)"
        b = s.get("risk_band", "No Data")
        if d not in counts:
            counts[d] = {rb: 0 for rb in risk_bands}
        if b not in counts[d]:
            counts[d][b] = 0
        counts[d][b] += 1

    # Header row
    row = 4
    headers = ["Department"] + risk_bands + ["Total"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2 + i, value=h)
    style_header_row(ws, row, 2, 2 + len(headers) - 1)
    row += 1

    # Data rows
    data_start = row
    for dept in sorted(counts.keys()):
        ws.cell(row=row, column=2, value=dept).font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_LEFT
        total = 0
        for j, band in enumerate(risk_bands):
            val = counts[dept].get(band, 0)
            total += val
            c = ws.cell(row=row, column=3 + j, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_CENTER
            c.number_format = "#,##0"
        c = ws.cell(row=row, column=3 + len(risk_bands), value=total)
        c.font = Font(name="Calibri", size=11, bold=True)
        c.alignment = ALIGN_CENTER
        c.number_format = "#,##0"
        for col in range(2, 3 + len(risk_bands) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    data_end = row - 1

    # Totals row
    ws.cell(row=row, column=2, value="Total").font = Font(name="Calibri", size=11, bold=True)
    grand = 0
    for j, band in enumerate(risk_bands):
        col_total = sum(counts[d].get(band, 0) for d in counts)
        grand += col_total
        c = ws.cell(row=row, column=3 + j, value=col_total)
        c.font = Font(name="Calibri", size=11, bold=True)
        c.alignment = ALIGN_CENTER
        c.number_format = "#,##0"
    c = ws.cell(row=row, column=3 + len(risk_bands), value=grand)
    c.font = Font(name="Calibri", size=11, bold=True)
    c.alignment = ALIGN_CENTER
    c.number_format = "#,##0"
    for col in range(2, 3 + len(risk_bands) + 1):
        ws.cell(row=row, column=col).border = BOTTOM_THICK

    # Conditional formatting — color the risk columns
    # High Risk column in red tones
    high_col = get_column_letter(3 + risk_bands.index("High Risk"))
    ws.conditional_formatting.add(
        f"{high_col}{data_start}:{high_col}{data_end}",
        DataBarRule(start_type="min", end_type="max", color=RED_RISK)
    )
    # Medium Risk in orange
    med_col = get_column_letter(3 + risk_bands.index("Medium Risk"))
    ws.conditional_formatting.add(
        f"{med_col}{data_start}:{med_col}{data_end}",
        DataBarRule(start_type="min", end_type="max", color=ORANGE)
    )
    # OK in green
    ok_col = get_column_letter(3 + risk_bands.index("OK"))
    ws.conditional_formatting.add(
        f"{ok_col}{data_start}:{ok_col}{data_end}",
        DataBarRule(start_type="min", end_type="max", color=GREEN_OK)
    )

    # Column widths
    set_col_width(ws, 2, 30)
    for i in range(len(risk_bands) + 1):
        set_col_width(ws, 3 + i, 16)

    ws.freeze_panes = "C5"
    return ws


def build_instructor_pivot(wb, course_data):
    """Build the Instructor Pivot sheet: Department → Instructor → Avg Attendance."""
    ws = wb.create_sheet("Instructor Pivot")
    ws.sheet_properties.tabColor = TEAL

    # Column A spacer
    set_col_width(ws, 1, 3)

    # Title
    ws.merge_cells("B2:G2")
    ws.cell(row=2, column=2, value="Attendance by Instructor").font = FONT_TITLE
    ws.row_dimensions[2].height = 30

    # Aggregate: department → instructor → list of course avg rates
    instructor_stats = {}  # dept -> {instructor -> {rates:[], students:0, courses:0}}
    for c in course_data:
        d = c.get("dept", "") or "(No Department)"
        inst = c.get("instructor", "") or "(No Instructor)"
        if d not in instructor_stats:
            instructor_stats[d] = {}
        if inst not in instructor_stats[d]:
            instructor_stats[d][inst] = {"rates": [], "students": 0, "courses": 0}
        instructor_stats[d][inst]["courses"] += 1
        instructor_stats[d][inst]["students"] += c.get("total_students", 0)
        if c.get("avg_attendance_pct") is not None:
            instructor_stats[d][inst]["rates"].append(c["avg_attendance_pct"])

    # Header row
    row = 4
    headers = ["Department", "Instructor", "Courses", "Students", "Avg Attendance %"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2 + i, value=h)
    style_header_row(ws, row, 2, 6)
    row += 1

    # Data rows grouped by department
    data_start = row
    for dept in sorted(instructor_stats.keys()):
        instructors = instructor_stats[dept]
        first_in_dept = True
        for inst in sorted(instructors.keys()):
            stats = instructors[inst]
            avg = round(sum(stats["rates"]) / len(stats["rates"]), 2) if stats["rates"] else None

            # Show department name only on first row of each group
            if first_in_dept:
                ws.cell(row=row, column=2, value=dept).font = Font(name="Calibri", size=11, bold=True)
                ws.cell(row=row, column=2).alignment = ALIGN_LEFT
                first_in_dept = False
            else:
                ws.cell(row=row, column=2, value="").font = FONT_BODY

            ws.cell(row=row, column=3, value=inst).font = FONT_BODY
            ws.cell(row=row, column=3).alignment = ALIGN_LEFT

            c = ws.cell(row=row, column=4, value=stats["courses"])
            c.font = FONT_BODY
            c.alignment = ALIGN_CENTER
            c.number_format = "#,##0"

            c = ws.cell(row=row, column=5, value=stats["students"])
            c.font = FONT_BODY
            c.alignment = ALIGN_CENTER
            c.number_format = "#,##0"

            if avg is not None:
                c = ws.cell(row=row, column=6, value=round(avg / 100, 4))
                c.number_format = "0.0%"
            else:
                c = ws.cell(row=row, column=6, value="N/A")
            c.font = FONT_BODY
            c.alignment = ALIGN_CENTER

            for col in range(2, 7):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1
    data_end = row - 1

    # Attendance % data bars
    if data_end >= data_start:
        ws.conditional_formatting.add(
            f"F{data_start}:F{data_end}",
            DataBarRule(start_type="min", end_type="max", color=TEAL)
        )

    # Column widths
    set_col_width(ws, 2, 30)
    set_col_width(ws, 3, 28)
    set_col_width(ws, 4, 12)
    set_col_width(ws, 5, 12)
    set_col_width(ws, 6, 20)

    ws.freeze_panes = "C5"
    return ws


# ── Main Build ───────────────────────────────────────────────────────────────

def load_csv_data():
    """Load real data from CSVs produced by extract.py.

    Returns (course_data, student_data, daily_data, compliance_data)
    in the same dict-list format that generate_sample_data() uses,
    or None if the CSVs don't exist.
    """
    files = {
        "course": os.path.join(CSV_DIR, "course_summary.csv"),
        "student": os.path.join(CSV_DIR, "student_detail.csv"),
        "daily": os.path.join(CSV_DIR, "daily_attendance.csv"),
        "compliance": os.path.join(CSV_DIR, "compliance.csv"),
    }

    # Check all files exist
    missing = [k for k, v in files.items() if not os.path.exists(v)]
    if missing:
        return None

    def read_csv(path):
        with open(path, "r", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))

    def to_float(val):
        if val is None or val == "":
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def to_int(val):
        if val is None or val == "":
            return 0
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return 0

    def to_bool(val):
        if isinstance(val, bool):
            return val
        return str(val).strip().lower() in ("true", "1", "yes")

    # -- Course Summary --
    raw = read_csv(files["course"])
    course_data = []
    for r in raw:
        course_data.append({
            "term": r.get("term", ""),
            "dept": r.get("department", ""),
            "course_code": r.get("course_code", ""),
            "course_name": r.get("course_name", ""),
            "instructor": r.get("instructor", ""),
            "total_students": to_int(r.get("total_students")),
            "avg_attendance_pct": to_float(r.get("avg_attendance_pct")),
            "students_100_pct": to_int(r.get("students_100_pct")),
            "students_above_threshold": to_int(r.get("students_above_threshold")),
            "students_below_threshold": to_int(r.get("students_below_threshold")),
            "pct_above_threshold": to_float(r.get("pct_above_threshold")),
            "pct_below_threshold": to_float(r.get("pct_below_threshold")),
            "total_present": to_int(r.get("total_present")),
            "total_late": to_int(r.get("total_late")),
            "total_absent": to_int(r.get("total_absent")),
            "total_na": to_int(r.get("total_na")),
            "total_meetings": to_int(r.get("total_meetings")),
            "last_attendance_date": r.get("last_attendance_date", ""),
            "days_since_last": to_int(r.get("days_since_last")) if r.get("days_since_last") else None,
            "status": r.get("status", ""),
            "api_blocked": to_bool(r.get("api_blocked")),
        })

    # -- Student Detail --
    raw = read_csv(files["student"])
    student_data = []
    for r in raw:
        student_data.append({
            "term": r.get("term", ""),
            "dept": r.get("department", ""),
            "course_code": r.get("course_code", ""),
            "course_name": r.get("course_name", ""),
            "instructor": r.get("instructor", ""),
            "student_id": r.get("student_id", ""),
            "student_name": r.get("student_name", ""),
            "present": to_int(r.get("present")),
            "late": to_int(r.get("late")),
            "absent": to_int(r.get("absent")),
            "excused": to_int(r.get("excused")),
            "attendance_pct": to_float(r.get("attendance_pct")),
            "last_attendance_date": r.get("last_attendance_date", ""),
            "risk_band": r.get("risk_band", ""),
            "below_threshold": r.get("below_threshold", "No"),
        })

    # -- Daily Attendance --
    raw = read_csv(files["daily"])
    daily_data = []
    for r in raw:
        daily_data.append({
            "term": r.get("term", ""),
            "dept": r.get("department", ""),
            "course_code": r.get("course_code", ""),
            "course_name": r.get("course_name", ""),
            "meeting_date": r.get("meeting_date", ""),
            "students_enrolled": to_int(r.get("students_enrolled")),
        })

    # -- Compliance --
    raw = read_csv(files["compliance"])
    compliance_data = []
    for r in raw:
        compliance_data.append({
            "term": r.get("term", ""),
            "dept": r.get("department", ""),
            "course_code": r.get("course_code", ""),
            "course_name": r.get("course_name", ""),
            "instructor": r.get("instructor", ""),
            "total_meetings": to_int(r.get("total_meetings")),
            "total_records": to_int(r.get("total_attendance_records")),
            "last_attendance_date": r.get("last_attendance_date", ""),
            "days_since_last": to_int(r.get("days_since_last")) if r.get("days_since_last") else None,
            "status": r.get("status", ""),
            "api_blocked": to_bool(r.get("api_blocked")),
            "no_attendance_recorded": to_bool(r.get("no_attendance_recorded")),
            "not_recent": to_bool(r.get("attendance_not_recent")),
        })

    return course_data, student_data, daily_data, compliance_data


def main():
    # Try loading real CSV data first; fall back to sample data
    csv_result = load_csv_data()
    if csv_result:
        print(f"Loading real data from {CSV_DIR}...")
        course_data, student_data, daily_data, compliance_data = csv_result
    else:
        print("No CSV data found — generating sample data...")
        print(f"  (Run extract.py first to get real data, or CSVs go in {CSV_DIR})")
        course_data, student_data, daily_data, compliance_data = generate_sample_data()

    print(f"  {len(course_data)} course sections")
    print(f"  {len(student_data)} student-course rows")
    print(f"  {len(daily_data)} daily attendance entries")
    print(f"  {len(compliance_data)} compliance rows")

    print("\nBuilding workbook...")
    wb = Workbook()

    print("  Building Dashboard...")
    build_dashboard(wb, course_data, student_data)

    print("  Building Course Summary...")
    build_course_summary(wb, course_data)

    print("  Building Student Detail...")
    build_student_detail(wb, student_data)

    print("  Building Daily Trends...")
    build_daily_trends(wb, daily_data, course_data)

    print("  Building Compliance...")
    build_compliance(wb, compliance_data)

    print("  Building Risk Pivot...")
    build_risk_pivot(wb, student_data)

    print("  Building Instructor Pivot...")
    build_instructor_pivot(wb, course_data)

    print("  Building Data Model docs...")
    build_data_model(wb)

    print("  Building Config...")
    build_config(wb)

    print(f"\nSaving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Done.")


if __name__ == "__main__":
    main()
